import json
import base64
from typing import Optional
import plivo
from plivo import plivoxml
import websockets
from fastapi import FastAPI, WebSocket, Request, Form, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.websockets import WebSocketDisconnect
import asyncio

from database.models import call_session_to_dict, transcript_entry_to_dict
from settings import settings
import uvicorn
import warnings
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime, timedelta
import re

# MongoDB imports
from database.db_service import db_service
from database.websocket_manager import websocket_manager

warnings.filterwarnings("ignore")
from dotenv import load_dotenv

load_dotenv()
records = []
p_index = 0
current_calling_customer = None  # Track the customer being called

# Global variable to store conversation transcripts
conversation_transcript = []

# Global variable to store current call session
current_call_session = None

plivo_client = plivo.RestClient(settings.PLIVO_AUTH_ID, settings.PLIVO_AUTH_TOKEN)

# Configuration
OPENAI_API_KEY = settings.AZURE_OPENAI_API_KEY_P
OPENAI_API_ENDPOINT = settings.AZURE_OPENAI_API_ENDPOINT_P
SYSTEM_MESSAGE = (
    "You are a helpful automotive service assistant"
)
VOICE = settings.DEFAULT_VOICE
LOG_EVENT_TYPES = [
    'error', 'response.content.done', 'rate_limits.updated',
    'response.done', 'input_audio_buffer.committed',
    'input_audio_buffer.speech_stopped', 'input_audio_buffer.speech_started',
    'session.created', 'conversation.item.input_audio_transcription.completed'
]
SHOW_TIMING_MATH = False
app = FastAPI()

not_registered_user_msg = "Sorry, we couldn't find your registered number. If you need any assistance, feel free to reach out. Thank you for calling, and have a great day!"

if not OPENAI_API_KEY:
    raise ValueError('Missing the OpenAI API key. Please set it in the .env file.')


def read_customer_records(filename=None):
    """Read customer records with automotive data"""
    global records
    records = []

    if filename is None:
        filename = settings.CUSTOMER_RECORDS_FILE

    if not os.path.exists(filename):
        print(f"⚠️ Customer records file '{filename}' not found. Please run generate_sample_data.py first.")
        return

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue

        record = {
            "name": row[0],
            "phone_number": row[1],
            "address": row[2],
            "car_model": row[3],
            "car_delivery_date": row[4],
            "last_servicing_date": row[5] if len(row) > 5 and row[5] else None,
        }
        records.append(record)

    print(f"✅ Loaded {len(records)} customer records from {filename}")


def determine_service_type(record):
    """Determine if customer needs 1st or 2nd servicing"""
    today = datetime.now().date()

    # Parse delivery date
    if isinstance(record["car_delivery_date"], str):
        try:
            delivery_date = datetime.strptime(record["car_delivery_date"], "%Y-%m-%d").date()
        except ValueError:
            print(f"⚠️ Invalid delivery date format for {record['name']}: {record['car_delivery_date']}")
            return None
    else:
        # If it's already a datetime object, convert to date
        delivery_date = record["car_delivery_date"].date() if isinstance(record["car_delivery_date"], datetime) else \
            record["car_delivery_date"]

    # Calculate days since delivery
    days_since_delivery = (today - delivery_date).days

    # If no last servicing date, check if 30+ days since delivery for 1st service
    if not record["last_servicing_date"]:
        if days_since_delivery >= settings.SERVICE_REMINDER_DAYS:
            return "first_service"
    else:
        # Parse last servicing date
        if isinstance(record["last_servicing_date"], str):
            try:
                last_service_date = datetime.strptime(record["last_servicing_date"], "%Y-%m-%d").date()
            except ValueError:
                print(f"⚠️ Invalid last service date format for {record['name']}: {record['last_servicing_date']}")
                return None
        else:
            # If it's already a datetime object, convert to date
            last_service_date = record["last_servicing_date"].date() if isinstance(record["last_servicing_date"],
                                                                                   datetime) else record[
                "last_servicing_date"]

        # Check if 9+ months since last service
        months_since_service = (today - last_service_date).days / 30.44  # Average days per month
        if months_since_service >= settings.REGULAR_SERVICE_MONTHS:
            return "second_service"

    return None


def get_eligible_customers():
    """Get list of customers eligible for service calls"""
    eligible_customers = []

    for i, record in enumerate(records):
        service_type = determine_service_type(record)
        if service_type:
            eligible_customers.append({
                "index": i,
                "record": record,
                "service_type": service_type
            })

    return eligible_customers


def get_current_customer_info():
    """Get current customer being called with proper indexing"""
    global current_calling_customer

    if current_calling_customer:
        return current_calling_customer

    eligible_customers = get_eligible_customers()

    # Current customer is the one at p_index (0-based for first call)
    current_index = p_index

    if current_index < len(eligible_customers):
        current_calling_customer = {
            "customer_record": eligible_customers[current_index]['record'],
            "service_type": eligible_customers[current_index]['service_type']
        }
        print(
            f"🎯 Current customer: {current_calling_customer['customer_record']['name']} - {current_calling_customer['service_type']}")
        return current_calling_customer
    else:
        print(f"⚠️ No customer found at index {current_index}")
        return None


def extract_appointment_details_from_response(confirmation_transcript):
    """
    Extract date and time information from the specific AI response that contains confirmation

    Args:
        confirmation_transcript (str): The specific AI response containing "बुक कर दी है"

    Returns:
        dict: Extracted appointment details from that specific response only
    """
    print(f"🔍 Extracting from confirmation response: {confirmation_transcript}")

    extracted_info = {
        "appointment_date": None,
        "appointment_time": None,
        "time_slot": None,
        "service_type": None,
        "confirmation_transcript": confirmation_transcript,
        "appointment_confirmed": True
    }

    # Enhanced date patterns - looking in the specific confirmation response
    date_patterns = [
        r'(\d{1,2}[-/]\d{1,2}[-/]\d{4})',  # DD-MM-YYYY or DD/MM/YYYY (21-06-2025)
        r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',  # YYYY-MM-DD or YYYY/MM/DD
        r'(\d{1,2}\s*\w+\s*\d{4})',  # DD Month YYYY (21 June 2025)
        r'(\d{1,2}\s*\w+)',  # DD Month (current year assumed)
    ]

    # Enhanced time slot patterns - looking in the specific confirmation response
    time_patterns = [
        r'(सुबह\s*\d{1,2}:\d{2})',  # सुबह 10:00
        r'(दोपहर\s*\d{1,2}:\d{2})',  # दोपहर 2:00
        r'(शाम\s*\d{1,2}:\d{2})',  # शाम 4:00
        r'(सुबह)',  # Morning
        r'(दोपहर)',  # Afternoon
        r'(शाम)',  # Evening
        r'(रात)',  # Night
        r'(\d{1,2}:\d{2})',  # HH:MM format
        r'(\d{1,2}\s*बजे)',  # X o'clock in Hindi
        r'(\d{1,2}\s*AM)',  # 10 AM
        r'(\d{1,2}\s*PM)',  # 2 PM
    ]

    # Extract dates from the confirmation response only
    for pattern in date_patterns:
        matches = re.findall(pattern, confirmation_transcript, re.IGNORECASE)
        if matches:
            extracted_info["appointment_date"] = matches[0]  # Get the first (and likely only) date
            print(f"📅 Found date in confirmation: {extracted_info['appointment_date']}")
            break

    # Extract time information from the confirmation response only
    for pattern in time_patterns:
        matches = re.findall(pattern, confirmation_transcript, re.IGNORECASE)
        if matches:
            extracted_info["appointment_time"] = matches[0]  # Get the first (and likely only) time
            print(f"⏰ Found time in confirmation: {extracted_info['appointment_time']}")
            break

    # Determine time slot from Hindi words in the confirmation response
    if 'सुबह' in confirmation_transcript:
        extracted_info["time_slot"] = "सुबह (Morning)"
    elif 'दोपहर' in confirmation_transcript:
        extracted_info["time_slot"] = "दोपहर (Afternoon)"
    elif 'शाम' in confirmation_transcript:
        extracted_info["time_slot"] = "शाम (Evening)"
    elif 'रात' in confirmation_transcript:
        extracted_info["time_slot"] = "रात (Night)"

    # If no specific time found, use time slot
    if not extracted_info["appointment_time"] and extracted_info["time_slot"]:
        extracted_info["appointment_time"] = extracted_info["time_slot"]

    # Determine service type from current customer
    current_customer_info = get_current_customer_info()
    if current_customer_info:
        extracted_info["service_type"] = current_customer_info['service_type']

    print(f"📊 Final extracted info from confirmation: {extracted_info}")
    return extracted_info


def append_appointment_to_excel(appointment_details, customer_record, filename=None):
    """
    Append appointment details to Excel file - Simplified version for automotive service

    Args:
        appointment_details (dict): Dictionary containing appointment info
        customer_record (dict): Dictionary containing customer info
        filename (str): Excel filename to write to
    """
    if filename is None:
        filename = settings.SERVICE_APPOINTMENTS_FILE

    headers = [
        "Customer Name",
        "Phone Number",
        "Car Model",
        "Service Type",
        "Appointment Date",
        "Appointment Time",
        "Address",
         # Changed from "Conversation Extract" to be more specific
    ]

    try:
        # Check if file exists
        if os.path.exists(filename):
            # Load existing workbook - THIS PRESERVES ALL EXISTING DATA
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            print(f"📊 Loaded existing Excel file with {ws.max_row} rows of data")
        else:
            # Create new workbook with headers ONLY if file doesn't exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Service Appointments"

            # Add headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            print("📊 Created new Excel file with headers")

        # Find the next empty row - THIS ENSURES NO OVERWRITING
        next_row = ws.max_row + 1
        print(f"📝 Appending data to row {next_row}")

        # Prepare service type display
        service_type_display = "First Service" if appointment_details.get(
            'service_type') == "first_service" else "Regular Service"

        # Get the specific confirmation transcript (not the entire conversation)
        confirmation_response = appointment_details.get('confirmation_transcript',
                                                        'No confirmation transcript available')

        # Prepare data row
        appointment_data = [
            customer_record.get('name', 'Unknown'),
            customer_record.get('phone_number', 'Unknown'),
            customer_record.get('car_model', 'Unknown'),
            service_type_display,
            appointment_details.get('appointment_date', 'Date to be confirmed'),
            appointment_details.get('appointment_time', 'Time to be confirmed'),
            customer_record.get('address', 'Unknown'),
        ]

        # Add data to the next row
        for col, value in enumerate(appointment_data, 1):
            ws.cell(row=next_row, column=col, value=str(value))

        # Save the workbook
        wb.save(filename)
        print(f"✅ Service appointment details saved to {filename} at row {next_row}")
        print(
            f"📋 Saved data: {customer_record.get('name')} - {appointment_details.get('appointment_date', 'TBC')} - {appointment_details.get('appointment_time', 'TBC')}")
        return True

    except Exception as e:
        print(f"❌ Error saving service appointment details: {e}")
        import traceback
        print(f"🔍 Traceback: {traceback.format_exc()}")
        return False


@app.get("/", response_class=JSONResponse)
async def index_page():
    return {"message": f"{settings.SERVICE_CENTER_NAME} Service Voice Agent is running!"}


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard():
    """Serve the automotive service dashboard"""
    try:
        with open("automotive_dashboard.html", "r", encoding="utf-8") as file:
            return HTMLResponse(content=file.read())
    except FileNotFoundError:
        return HTMLResponse(
            content="<h1>Dashboard not found</h1><p>Please ensure automotive_dashboard.html exists in the project directory.</p>",
            status_code=404
        )


@app.websocket("/ws/transcripts")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time transcript updates"""
    await websocket_manager.connect(websocket, connection_type="dashboard")
    try:
        # Send initial connection confirmation
        await websocket.send_text(json.dumps({
            "type": "connection_status",
            "status": "connected",
            "timestamp": datetime.utcnow().isoformat()
        }))

        while True:
            try:
                # Set a timeout to prevent indefinite blocking
                message = await asyncio.wait_for(
                    websocket.receive_text(),
                    timeout=30.0
                )

                # Parse and handle incoming messages
                try:
                    data = json.loads(message)

                    # Handle ping messages
                    if data.get("type") == "ping":
                        await websocket.send_text(json.dumps({
                            "type": "pong",
                            "timestamp": datetime.utcnow().isoformat()
                        }))

                    # Handle other message types as needed
                    print(f"📱 Received from dashboard: {data}")

                except json.JSONDecodeError:
                    print(f"⚠️ Invalid JSON received: {message}")

            except asyncio.TimeoutError:
                # Send keepalive ping
                try:
                    await websocket.send_text(json.dumps({
                        "type": "keepalive",
                        "timestamp": datetime.utcnow().isoformat()
                    }))
                except:
                    break  # Connection is broken

    except WebSocketDisconnect:
        print("📱 Dashboard WebSocket disconnected")
    except Exception as e:
        print(f"❌ WebSocket error: {e}")
    finally:
        websocket_manager.disconnect(websocket)


@app.get("/appointment-details")
async def get_appointment_details():
    """API endpoint to get extracted appointment details - Note: Now uses response-specific extraction"""
    # This endpoint would need to be updated to work with the new response-specific extraction
    # For now, return a message indicating the change
    return JSONResponse({
        "message": "Appointment extraction now happens automatically when 'बुक कर दी है' is detected in AI responses",
        "method": "Response-specific extraction (not full conversation)"
    })


@app.get("/eligible-customers")
async def get_eligible_customers_api():
    """API endpoint to get customers eligible for service"""
    eligible = get_eligible_customers()
    return JSONResponse(eligible)


@app.get("/api/recent-calls")
async def get_recent_calls():
    """Get recent call sessions"""
    try:
        recent_calls = await db_service.get_recent_calls(limit=20)
        return [call_session_to_dict(call) for call in recent_calls]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/call-transcripts/{call_id}")
async def get_call_transcripts(call_id: str):
    """Get transcripts for a specific call"""
    try:
        transcripts = await db_service.get_call_transcripts(call_id)
        return [transcript_entry_to_dict(transcript) for transcript in transcripts]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/service-statistics")
async def get_service_statistics():
    """Get automotive service statistics"""
    try:
        stats = await db_service.get_call_statistics()
        return stats
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/calls-by-car-model/{car_model}")
async def get_calls_by_car_model(car_model: str):
    """Get calls for specific car model"""
    try:
        calls = await db_service.get_calls_by_car_model(car_model)
        return [call_session_to_dict(call) for call in calls]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/calls-by-service-type/{service_type}")
async def get_calls_by_service_type(service_type: str):
    """Get calls for specific service type"""
    try:
        calls = await db_service.get_calls_by_service_type(service_type)
        return [call_session_to_dict(call) for call in calls]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/customer-history/{phone_number}")
async def get_customer_history(phone_number: str):
    """Get call history for specific customer"""
    try:
        calls = await db_service.get_calls_by_phone(phone_number)
        return [call_session_to_dict(call) for call in calls]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/active-calls")
async def get_active_calls():
    """Get currently active calls - Note: No status tracking in DB anymore"""
    try:
        # Since we don't track status in DB, return recent calls from today
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        calls = []
        recent_calls = await db_service.get_recent_calls(limit=50)

        # Filter calls from today (assuming they might still be active)
        for call in recent_calls:
            if call.started_at >= today_start:
                calls.append(call)

        return [call_session_to_dict(call) for call in calls]
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.api_route("/webhook", methods=["GET", "POST"])
def home(request: Request):
    """Handle webhook for making calls to next eligible customer"""
    global p_index, current_calling_customer
    if request.method == "POST":
        eligible_customers = get_eligible_customers()

        if p_index < len(eligible_customers):
            current_customer = eligible_customers[p_index]
            # Set the current calling customer BEFORE making the call
            current_calling_customer = {
                "customer_record": current_customer['record'],
                "service_type": current_customer['service_type']
            }

            call_made = plivo_client.calls.create(
                from_=settings.PLIVO_FROM_NUMBER,
                to_=current_customer['record']['phone_number'],
                answer_url=settings.PLIVO_ANSWER_XML,
                answer_method='GET')

            print(f"📞 Webhook POST request detected! Calling {current_customer['record']['name']} (Index: {p_index})")
            p_index += 1
        else:
            print("⚠️ No more eligible customers to call")

    xml_data = f'''<?xml version="1.0" encoding="UTF-8"?>
    <Response>
        <Speak>Please wait while we connect your call to the {settings.SERVICE_CENTER_NAME} AI Agent. OK you can start speaking.</Speak>
        <Stream streamTimeout="86400" keepCallAlive="true" bidirectional="true" contentType="audio/x-mulaw;rate=8000" audioTrack="inbound" >
            {settings.HOST_URL}/media-stream
        </Stream>
    </Response>
    '''
    return HTMLResponse(xml_data, media_type='application/xml')


@app.api_route("/incoming-call", methods=["GET", "POST"])
async def handle_incoming_call(request: Request):
    """Handle incoming call and return TwiML response to connect to Media Stream"""
    form_data = await request.form()
    caller_phone = form_data.get("From", "unknown")
    request.state.caller_phone = caller_phone

    wss_host = settings.HOST_URL
    http_host = wss_host.replace('wss://', 'https://')

    response = plivoxml.ResponseElement()

    get_input = plivoxml.GetInputElement() \
        .set_action(f"{http_host}/voice") \
        .set_method("POST") \
        .set_input_type("dtmf") \
        .set_redirect(True) \
        .set_language(settings.SECONDARY_LANGUAGE) \
        .set_num_digits(1)

    get_input.add_speak(
        content="To switch to Hindi, please press 5. To continue in English, press any other key.",
        voice="Polly.Salli",
        language=settings.SECONDARY_LANGUAGE
    )

    response.add(get_input)
    response.add_speak(
        content="No selection received. Continuing in English.",
        voice="Polly.Salli",
        language=settings.SECONDARY_LANGUAGE
    )

    return HTMLResponse('<?xml version="1.0" encoding="UTF-8"?>\n' + response.to_string(), media_type="application/xml")


@app.post("/voice")
async def voice_post(Digits: Optional[str] = Form(None)):
    """Handle the user's input"""
    response = plivoxml.ResponseElement()
    lang_code = settings.SECONDARY_LANGUAGE

    if Digits == '5':
        lang_code = 'hi-IN'
        response.add(plivoxml.SpeakElement('नमस्ते, मैं आपकी कैसे मदद कर सकती हूँ?', language=lang_code))
    else:
        response.add(plivoxml.SpeakElement('Hello, How can I help you today?', language=lang_code))

    wss_host = settings.HOST_URL

    stream = response.add(plivoxml.StreamElement(f'{wss_host}/media-stream', extraHeaders=f"lang_code={lang_code}",
                                                 bidirectional=True,
                                                 streamTimeout=86400,
                                                 keepCallAlive=True,
                                                 contentType="audio/x-mulaw;rate=8000",
                                                 audioTrack="inbound"
                                                 ))

    return HTMLResponse('<?xml version="1.0" encoding="UTF-8"?>\n' + stream.to_string(), media_type="application/xml")


@app.websocket("/media-stream")
async def handle_media_stream(websocket: WebSocket):
    """Handle WebSocket connections between Plivo and OpenAI"""
    global conversation_transcript, current_call_session, current_calling_customer

    await websocket.accept()

    # Get current customer info using the proper method
    current_customer_info = get_current_customer_info()

    if current_customer_info:
        customer_record = current_customer_info['customer_record']
        service_type = current_customer_info['service_type']
        print(f"🎯 WebSocket: Using customer {customer_record['name']} with service type {service_type}")
    else:
        # Fallback for unknown customer
        customer_record = {"name": "Unknown Customer", "phone_number": "Unknown", "car_model": "Unknown"}
        service_type = None
        print("⚠️ WebSocket: Using fallback customer data")

    # Create new call session in MongoDB - Updated field names
    current_call_session = await db_service.create_call_session(
        customer_name=customer_record.get("name", "Unknown Customer"),  # Changed from patient_name
        customer_phone=customer_record.get("phone_number", "Unknown"),  # Changed from patient_phone
        car_model=customer_record.get("car_model"),
        service_type=service_type
    )

    # Broadcast call started status - Updated to use customer fields
    await websocket_manager.broadcast_call_status(
        call_id=current_call_session.call_id,
        status="started",
        patient_name=current_call_session.customer_name,  # Uses customer_name now
        car_model=customer_record.get("car_model"),
        service_type=service_type,
        phone_number=current_call_session.customer_phone  # Uses customer_phone now
    )

    # Broadcast customer info to dashboard
    await websocket_manager.broadcast_customer_info(
        call_id=current_call_session.call_id,
        customer_data=customer_record
    )

    user_details = None

    async with websockets.connect(
            OPENAI_API_ENDPOINT,
            extra_headers={"api-key": OPENAI_API_KEY},
            ping_timeout=20,
            close_timeout=10
    ) as realtime_ai_ws:
        await initialize_session(realtime_ai_ws, user_details)

        stream_sid = None
        latest_media_timestamp = 0
        last_assistant_item = None
        mark_queue = []
        response_start_timestamp_twilio = None

        async def receive_from_twilio():
            nonlocal stream_sid, latest_media_timestamp
            try:
                async for message in websocket.iter_text():
                    data = json.loads(message)
                    if data['event'] == 'media' and realtime_ai_ws.open:
                        latest_media_timestamp = int(data['media']['timestamp'])
                        audio_append = {
                            "type": "input_audio_buffer.append",
                            "audio": data['media']['payload']
                        }
                        await realtime_ai_ws.send(json.dumps(audio_append))
                    elif data['event'] == 'start':
                        stream_sid = data['start']['streamId']
                        print(f"📞 Incoming stream has started {stream_sid}")
                        await realtime_ai_ws.send(json.dumps(data))
                        response_start_timestamp_twilio = None
                        latest_media_timestamp = 0
                        last_assistant_item = None
                    elif data['event'] == 'mark':
                        if mark_queue:
                            mark_queue.pop(0)
            except WebSocketDisconnect:
                print("📞 Client disconnected.")
                if realtime_ai_ws.open:
                    await realtime_ai_ws.close()

                # Broadcast status for UI purposes
                if current_call_session:
                    await websocket_manager.broadcast_call_status(
                        call_id=current_call_session.call_id,
                        status="ended",
                        car_model=customer_record.get("car_model"),
                        service_type=service_type
                    )

        async def send_to_twilio():
            nonlocal stream_sid, last_assistant_item, response_start_timestamp_twilio
            try:
                async for openai_message in realtime_ai_ws:
                    response = json.loads(openai_message)

                    # Handle user transcription - UNIFIED HANDLING
                    if response.get('type') == 'conversation.item.input_audio_transcription.completed':
                        try:
                            print(f"🎤 RAW TRANSCRIPTION RESPONSE: {response}")
                            user_transcript = response.get('transcript', '').strip()

                            if user_transcript:
                                print(f"👤 Customer said: {user_transcript}")

                                # Store user transcript in MongoDB and broadcast
                                if current_call_session:
                                    await db_service.save_transcript(
                                        call_id=current_call_session.call_id,
                                        speaker="user",
                                        message=user_transcript
                                    )

                                    # Broadcast to WebSocket clients
                                    await websocket_manager.broadcast_transcript(
                                        call_id=current_call_session.call_id,
                                        speaker="user",
                                        message=user_transcript,
                                        timestamp=datetime.utcnow().isoformat(),
                                        car_model=customer_record.get("car_model"),
                                        service_type=service_type
                                    )

                                # Add user transcript to global conversation for appointment detection
                                conversation_transcript.append(user_transcript)

                        except Exception as e:
                            print(f"❌ Error processing user transcript: {e}")

                    # Handle AI response transcription
                    elif response['type'] in LOG_EVENT_TYPES:
                        try:
                            transcript = response['response']['output'][0]['content'][0]['transcript']
                            print(f"🤖 AI Response: {transcript}")

                            # Store AI response in MongoDB and broadcast
                            if current_call_session:
                                await db_service.save_transcript(
                                    call_id=current_call_session.call_id,
                                    speaker="ai",
                                    message=transcript
                                )

                                # Broadcast to WebSocket clients
                                await websocket_manager.broadcast_transcript(
                                    call_id=current_call_session.call_id,
                                    speaker="ai",
                                    message=transcript,
                                    timestamp=datetime.utcnow().isoformat(),
                                    car_model=customer_record.get("car_model"),
                                    service_type=service_type
                                )

                            # Add AI transcript to global conversation for appointment detection
                            conversation_transcript.append(transcript)

                            # *** UPDATED APPOINTMENT DETECTION LOGIC ***
                            # Check specifically for appointment confirmation keyword in THIS SPECIFIC RESPONSE
                            if "बुक कर दी है" in transcript:
                                print(f"🎯 APPOINTMENT CONFIRMATION DETECTED: {transcript}")

                                # Extract appointment details ONLY from this specific confirmation response
                                current_details = extract_appointment_details_from_response(transcript)
                                print(f"📋 Extracted details from confirmation: {current_details}")

                                # Get current customer info
                                current_customer_info = get_current_customer_info()
                                if current_customer_info:
                                    current_customer_record = current_customer_info['customer_record']

                                    # Save to Excel using the new function
                                    success = append_appointment_to_excel(current_details, current_customer_record)

                                    if success:
                                        print(f"✅ APPOINTMENT SAVED TO EXCEL!")

                                        # Broadcast appointment confirmation
                                        await websocket_manager.broadcast_appointment_confirmation(
                                            call_id=current_call_session.call_id,
                                            customer_name=current_customer_record.get("name"),
                                            appointment_date=current_details.get("appointment_date", "To be confirmed"),
                                            appointment_time=current_details.get("appointment_time", "To be confirmed"),
                                            car_model=current_customer_record.get("car_model"),
                                            service_type=service_type or "Service"
                                        )
                                    else:
                                        print(f"❌ Failed to save appointment to Excel")
                                else:
                                    print(f"⚠️ No customer info available for Excel save")

                        except (KeyError, IndexError):
                            print("⚠️ No transcript found in response")

                    # Handle audio delta
                    elif response.get('type') == 'response.audio.delta' and 'delta' in response:
                        audio_payload = base64.b64encode(base64.b64decode(response['delta'])).decode('utf-8')
                        audio_delta = {
                            "event": "playAudio",
                            "media": {
                                "contentType": 'audio/x-mulaw',
                                "sampleRate": 8000,
                                "payload": audio_payload
                            }
                        }
                        await websocket.send_json(audio_delta)

                        if response_start_timestamp_twilio is None:
                            response_start_timestamp_twilio = latest_media_timestamp
                            if SHOW_TIMING_MATH:
                                print(
                                    f"⏱️ Setting start timestamp for new response: {response_start_timestamp_twilio}ms")

                        if response.get('item_id'):
                            last_assistant_item = response['item_id']

                        await send_mark(websocket, stream_sid)

                    # Handle speech started
                    elif response.get('type') == 'input_audio_buffer.speech_started':
                        print("🎙️ Speech started detected.")
                        if last_assistant_item:
                            print(f"⏸️ Interrupting response with id: {last_assistant_item}")
                            await handle_speech_started_event()
            except Exception as e:
                print(f"❌ Error in send_to_twilio: {e}")

        async def handle_speech_started_event():
            nonlocal response_start_timestamp_twilio, last_assistant_item
            print("🔄 Handling speech started event.")
            if mark_queue and response_start_timestamp_twilio is not None:
                elapsed_time = latest_media_timestamp - response_start_timestamp_twilio
                if SHOW_TIMING_MATH:
                    print(
                        f"⏱️ Calculating elapsed time for truncation: {latest_media_timestamp} - {response_start_timestamp_twilio} = {elapsed_time}ms")

                if last_assistant_item:
                    if SHOW_TIMING_MATH:
                        print(f"✂️ Truncating item with ID: {last_assistant_item}, Truncated at: {elapsed_time}ms")

                    truncate_event = {
                        "type": "conversation.item.truncate",
                        "item_id": last_assistant_item,
                        "content_index": 0,
                        "audio_end_ms": elapsed_time
                    }
                    await realtime_ai_ws.send(json.dumps(truncate_event))

                await websocket.send_json({
                    "event": "clear",
                    "streamSid": stream_sid
                })

                mark_queue.clear()
                last_assistant_item = None
                response_start_timestamp_twilio = None

        async def send_mark(connection, stream_sid):
            if stream_sid:
                mark_event = {
                    "event": "mark",
                    "streamSid": stream_sid,
                    "mark": {"name": "responsePart"}
                }
                await connection.send_json(mark_event)
                mark_queue.append('responsePart')

        await asyncio.gather(receive_from_twilio(), send_to_twilio())


async def send_initial_conversation_item(realtime_ai_ws, user_details=None):
    """Send initial conversation item with personalized greeting"""
    current_customer_info = get_current_customer_info()

    if current_customer_info:
        current_customer = current_customer_info['customer_record']
        greeting_name = current_customer.get("name", "Sir/Madam")
    else:
        greeting_name = "Sir/Madam"
        current_customer = {"name": "Customer", "car_model": ""}

    initial_conversation_item = {
        "type": "conversation.item.create",
        "item": {
            "type": "message",
            "role": "assistant",
            "content": [{
                "type": "text",
                "text": f"Hey {greeting_name}! I am calling from {settings.SERVICE_CENTER_NAME}."
            }]
        }
    }
    await realtime_ai_ws.send(json.dumps(initial_conversation_item))
    await realtime_ai_ws.send(json.dumps({"type": "response.create"}))


async def initialize_session(realtime_ai_ws, user_details=None):
    """Control initial session with OpenAI"""
    current_customer_info = get_current_customer_info()

    if current_customer_info:
        current_customer = current_customer_info['customer_record']
        service_type = current_customer_info['service_type']

        # Calculate service timing info
        today = datetime.now().date()
        if isinstance(current_customer["car_delivery_date"], str):
            delivery_date = datetime.strptime(current_customer["car_delivery_date"], "%Y-%m-%d").date()
        else:
            delivery_date = current_customer["car_delivery_date"]
            if isinstance(delivery_date, datetime):
                delivery_date = delivery_date.date()

        days_since_delivery = (today - delivery_date).days

        service_message = ""
        if service_type == "first_service":
            service_message = f"This is their first service call. Their car was delivered {days_since_delivery} days ago."
        else:
            if current_customer["last_servicing_date"]:
                if isinstance(current_customer["last_servicing_date"], str):
                    last_service = datetime.strptime(current_customer["last_servicing_date"], "%Y-%m-%d").date()
                else:
                    last_service = current_customer["last_servicing_date"]
                    if isinstance(last_service, datetime):
                        last_service = last_service.date()
                months_since_service = (today - last_service).days / 30.44
                service_message = f"This is a regular service reminder. Their last service was {months_since_service:.1f} months ago."
    else:
        current_customer = {"name": "Customer", "car_model": ""}
        service_message = "This is a general service inquiry."

    session_update = {
        "type": "session.update",
        "session": {
            "input_audio_transcription": {
                "model": "whisper-1",
                "language": settings.PRIMARY_LANGUAGE,
            },
            "turn_detection": {"type": "server_vad"},
            "input_audio_format": "g711_ulaw",
            "output_audio_format": "g711_ulaw",
            "voice": VOICE,
            "instructions": f'''AI ROLE: Female voice representative from automotive service center
LANGUAGE: Hindi (देवनागरी लिपि) 
VOICE STYLE: Professional, friendly, helpful, feminine
GENDER CONSISTENCY: Use feminine forms (e.g., "बोल रही हूँ", "कर सकती हूँ", "समझ सकती हूँ")
GOAL: Schedule car servicing appointment and handle customer responses

You are talking to {current_customer['name']}, who owns a {current_customer.get('car_model', 'car')}.
{service_message}

CONVERSATION FLOW:

"नमस्ते {current_customer['name']} जी, मैं Patni Toyota Nagpur से Priya बोल रही हूँ। आप कैसे हैं?"

(रुकें, उत्तर सुनें)

"मैं आपको यह बताने के लिए कॉल कर रही हूँ कि आपकी {current_customer.get('car_model', 'गाड़ी')} की सर्विसिंग का समय हो गया है। क्या आप अपॉइंटमेंट बुक कराना चाहेंगे?"

IF USER SAYS YES / INTERESTED:

"बहुत अच्छा! मैं आपको कुछ उपलब्ध तारीखें बताती हूँ —"

"क्या आप {(datetime.today() + timedelta(days=1)).strftime("%d-%m-%Y")}, {(datetime.today() + timedelta(days=2)).strftime("%d-%m-%Y")} को लाना पसंद करेंगे?"

(रुकें, तारीख चुनने दें)

"और उस दिन आपको कौन-सा समय ठीक लगेगा — सुबह, दोपहर या शाम?"

(रुकें, समय चुनने दें)

"शानदार! तो मैंने आपकी {current_customer.get('car_model', 'गाड़ी')} की सर्विसिंग {{चुनी हुई तारीख}} को {{चुना हुआ समय}} के लिए बुक कर दी है।"

IF USER SAYS NO / NOT NOW:

"कोई बात नहीं — जब भी आप तैयार हों, हमें कॉल कर सकते हैं। धन्यवाद!"

IMPORTANT NOTES:
- Be empathetic and understanding
- If customer has concerns about cost, mention competitive pricing
- If they ask about service details, explain basic maintenance check
- Always confirm appointment details clearly
- Keep conversation natural and friendly''',
            "modalities": ["text", "audio"],
            "temperature": 0.7,
        }
    }
    print('📤 Sending session update:', json.dumps(session_update))
    await realtime_ai_ws.send(json.dumps(session_update))

    await send_initial_conversation_item(realtime_ai_ws, user_details)


@app.on_event("startup")
async def startup_event():
    """Initialize database connection on startup"""
    connected = await db_service.connect()
    if not connected:
        raise RuntimeError("Failed to connect to MongoDB")
    print(f"✅ {settings.SERVICE_CENTER_NAME} Application started with MongoDB connection")

    # Start WebSocket manager periodic tasks
    await websocket_manager.start_periodic_tasks()


@app.on_event("shutdown")
async def shutdown_event():
    """Close database connection on shutdown"""
    await db_service.disconnect()
    print("👋 Application shutdown complete")


def make_next_call():
    """Make a call to the next eligible customer"""
    global p_index, current_calling_customer

    eligible_customers = get_eligible_customers()

    if p_index < len(eligible_customers):
        current_customer = eligible_customers[p_index]

        # Set the current calling customer BEFORE making the call
        current_calling_customer = {
            "customer_record": current_customer['record'],
            "service_type": current_customer['service_type']
        }

        try:
            call_made = plivo_client.calls.create(
                from_=settings.PLIVO_FROM_NUMBER,
                to_=current_customer['record']['phone_number'],
                answer_url=settings.PLIVO_ANSWER_XML,
                answer_method='GET'
            )

            service_display = "First Service" if current_customer[
                                                     'service_type'] == "first_service" else "Regular Service"
            print(f"📞 Called {current_customer['record']['name']} for {service_display} (Index: {p_index})")

            # Increment index for next call
            p_index += 1

            return True

        except Exception as e:
            print(f"❌ Failed to make call to {current_customer['record']['name']}: {e}")
            return False
    else:
        print("⚠️ No more eligible customers to call")
        return False


def main():
    global p_index, current_calling_customer

    print(f"🚗 Starting {settings.SERVICE_CENTER_NAME} AI Service Call System")
    print("=" * 60)

    # Reset global variables
    p_index = 0
    current_calling_customer = None

    # Read customer records
    read_customer_records()

    if not records:
        print("❌ No customer records found. Please run generate_sample_data.py first.")
        return

    # Get eligible customers for service
    eligible_customers = get_eligible_customers()
    print(f"📊 Found {len(eligible_customers)} customers eligible for service calls:")
    for i, customer in enumerate(eligible_customers):
        service_display = "First Service" if customer['service_type'] == "first_service" else "Regular Service"
        print(f"   {i + 1}. {customer['record']['name']}: {service_display} ({customer['record']['car_model']})")

    if eligible_customers:
        print(f"\n📞 Making first call to {eligible_customers[0]['record']['name']}...")
        success = make_next_call()
        if not success:
            print("❌ Failed to make initial call")
    else:
        print("⚠️ No customers eligible for service calls at this time")

    print(f"\n🌐 Starting server on http://localhost:{settings.PORT}")
    print(f"📊 Dashboard available at: http://localhost:{settings.PORT}/dashboard")
    print(f"🔗 API documentation at: http://localhost:{settings.PORT}/docs")
    print("\n🎯 System Ready!")

    uvicorn.run(app, host="0.0.0.0", port=settings.PORT)


if __name__ == "__main__":
    main()
