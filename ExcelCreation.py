# import openpyxl
# from openpyxl import Workbook
# from datetime import datetime, timedelta
# import random
#
#
# def generate_customer_records():
#     """Generate sample customer records for automotive service"""
#
#     # Sample data
#     first_names = [
#         "Rajesh", "Priya", "Amit", "Sunita", "Vikash", "Neha", "Suresh", "Kavita",
#         "Manoj", "Deepika", "Ravi", "Pooja", "Santosh", "Meera", "Ajay", "Shweta",
#         "Arjun", "Sneha", "Rohit", "Anjali", "Kiran", "Rekha", "Anil", "Geeta",
#         "Nitin", "Swati", "Vishal", "Rashmi", "Ramesh", "Nisha"
#     ]
#
#     last_names = [
#         "Sharma", "Patel", "Singh", "Kumar", "Gupta", "Verma", "Agarwal", "Joshi",
#         "Mehta", "Shah", "Yadav", "Mishra", "Tiwari", "Sinha", "Pandey", "Jain",
#         "Saxena", "Dubey", "Shukla", "Thakur", "Chopra", "Bansal", "Malhotra", "Arora"
#     ]
#
#     car_models = [
#         "Toyota Fortuner", "Toyota Innova Crysta", "Toyota Glanza", "Toyota Urban Cruiser",
#         "Toyota Camry", "Toyota Vellfire", "Toyota Yaris", "Toyota Etios", "Toyota Corolla Altis",
#         "Toyota Prius", "Toyota Land Cruiser", "Toyota Hiace", "Toyota Qualis"
#     ]
#
#     areas_nagpur = [
#         "Civil Lines", "Dharampeth", "Sadar", "Sitabuldi", "Ramdaspeth", "Bajaj Nagar",
#         "Laxmi Nagar", "Pratap Nagar", "Shankar Nagar", "Manish Nagar", "Trimurti Nagar",
#         "Gandhibagh", "Itwari", "Mahal", "Lakadganj", "Cotton Market", "Medical Square"
#     ]
#
#     # Create workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Customer Records"
#
#     # Headers
#     headers = [
#         "Name",
#         "Phone Number",
#         "Address",
#         "Car Model",
#         "Car Delivery Date",
#         "Last Servicing Date"
#     ]
#
#     for col, header in enumerate(headers, 1):
#         ws.cell(row=1, column=col, value=header)
#
#     # Generate sample customers
#     today = datetime.now().date()
#
#     for i in range(2, 32):  # Generate 30 customers
#         first_name = random.choice(first_names)
#         last_name = random.choice(last_names)
#         full_name = f"{first_name} {last_name}"
#
#         # Enter targeted number
#         phone = f"+91"
#
#         # Generate address
#         area = random.choice(areas_nagpur)
#         house_no = random.randint(1, 999)
#         address = f"{house_no}, {area}, Nagpur, Maharashtra"
#
#         # Generate car model
#         car_model = random.choice(car_models)
#
#         # Generate delivery date (random between 1 month to 2 years ago)
#         days_back = random.randint(30, 730)
#         delivery_date = today - timedelta(days=days_back)
#
#         # Generate last servicing date (some customers have it, some don't)
#         last_servicing = None
#         if random.random() > 0.3:  # 70% chance of having previous service
#             # Last service between 2-15 months ago
#             service_days_back = random.randint(60, 450)
#             last_servicing = today - timedelta(days=service_days_back)
#             # Make sure last service is after delivery
#             if last_servicing < delivery_date:
#                 last_servicing = delivery_date + timedelta(days=random.randint(30, 90))
#
#         # Add row to worksheet
#         ws.cell(row=i, column=1, value=full_name)
#         ws.cell(row=i, column=2, value=phone)
#         ws.cell(row=i, column=3, value=address)
#         ws.cell(row=i, column=4, value=car_model)
#         ws.cell(row=i, column=5, value=delivery_date.strftime("%Y-%m-%d"))
#         ws.cell(row=i, column=6, value=last_servicing.strftime("%Y-%m-%d") if last_servicing else "")
#
#     # Save the file
#     filename = "Customer_Records.xlsx"
#     wb.save(filename)
#     print(f"✅ Generated {filename} with 30 sample customer records")
#
#     # Print some statistics
#     print("\n📊 Customer Statistics:")
#     eligible_first = 0
#     eligible_regular = 0
#
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         delivery_date = datetime.strptime(row[4], "%Y-%m-%d").date()
#         last_service = datetime.strptime(row[5], "%Y-%m-%d").date() if row[5] else None
#
#         days_since_delivery = (today - delivery_date).days
#
#         if not last_service and days_since_delivery >= 30:
#             eligible_first += 1
#         elif last_service:
#             months_since_service = (today - last_service).days / 30.44
#             if months_since_service >= 9:
#                 eligible_regular += 1
#
#     print(f"- Eligible for First Service: {eligible_first}")
#     print(f"- Eligible for Regular Service: {eligible_regular}")
#     print(f"- Total Eligible for Service: {eligible_first + eligible_regular}")
#
#     return filename
#
#
# def generate_service_appointments_template():
#     """Generate template for service appointments"""
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Service Appointments"
#
#     headers = [
#         "Name",
#         "Phone Number",
#         "Car Model",
#         "Service Type",
#         "Appointment Date",
#         "Time Slot",
#         "Address",
#         "Car Delivery Date",
#         "Last Servicing Date",
#         "Booking Timestamp"
#     ]
#
#     for col, header in enumerate(headers, 1):
#         ws.cell(row=1, column=col, value=header)
#
#     filename = "Service_Appointments.xlsx"
#     wb.save(filename)
#     print(f"✅ Generated {filename} template for appointments")
#     return filename
#
#
# if __name__ == "__main__":
#     print("🚗 Generating Sample Data for Patni Toyota Service System")
#     print("=" * 60)
#
#     # Generate customer records
#     customer_file = generate_customer_records()
#
#     print()
#
#     # Generate appointments template
#     appointment_file = generate_service_appointments_template()
#
#     print("\n🎉 Sample data generation complete!")
#     print(f"📁 Files created:")
#     print(f"   - {customer_file}")
#     print(f"   - {appointment_file}")
#     print("\n💡 You can now run the automotive service main.py file!")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime



def create_appointments_file():
    """Create empty appointments file with headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Service Appointments"

    headers = [
        "Customer Name", "Phone Number", "Car Model", "Service Type",
        "Appointment Date", "Appointment Time", "Address"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    filename = "Service_Appointments.xlsx"
    wb.save(filename)
    print(f"✅ Created empty appointments file: {filename}")

create_appointments_file()
